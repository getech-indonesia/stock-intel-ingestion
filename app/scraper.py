import requests
import cloudscraper
import re
from io import BytesIO
from pypdf import PdfReader
from openpyxl import load_workbook
import xlrd
from config.settings import REQUEST_TIMEOUT

QUARTER_MAP = {
    "Q1": "tw1",
    "Q2": "tw2",
    "Q3": "tw3",
    "Q4": "tw4",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": "https://www.idx.co.id/",
}

BASE_URL = "https://www.idx.co.id"
SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}
MAX_ATTACHMENTS_TO_PARSE = 4
MAX_CHARS_PER_FILE = 8000
MAX_TOTAL_CHARS = 14000

PRIORITY_KEYWORDS = [
    "financial",
    "keuangan",
    "laporan-keuangan",
    "quarter",
    "kuartal",
    "interim",
]
LOW_PRIORITY_KEYWORDS = [
    "esg",
    "sustainability",
    "keberlanjutan",
]

FINANCIAL_TEXT_KEYWORDS = [
    "jumlah aset",
    "total assets",
    "jumlah liabilitas",
    "total liabilities",
    "jumlah ekuitas",
    "total equity",
    "pendapatan",
    "revenue",
    "laba operasional",
    "operating profit",
    "laba bersih",
    "net profit",
    "laba tahun berjalan",
    "profit for the period",
    "beban operasional",
    "operating expense",
    "net interest income",
    "current ratio",
    "debt to equity",
    "return on assets",
    "return on equity",
    "eps",
    "book value per share",
]

# Shareholder keywords — nama & angka sering di baris terpisah, perlu context window lebih lebar
SHAREHOLDER_TEXT_KEYWORDS = [
    "pemegang saham",
    "shareholder",
    "komposisi pemegang saham",
    "composition of shareholders",
    "jumlah lembar saham",
    "number of shares",
    "persentase kepemilikan",
    "percentage of ownership",
    "kepemilikan saham",
    "share ownership",
    "daftar pemegang",
]

NUMERIC_ROW_RE = re.compile(r"\d[\d,\.]*")


def _create_scraper():
    return cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )

def _get(url: str, params: dict) -> dict:
    """Helper function to make GET requests to IDX API"""
    try:
        # IDX often blocks plain requests clients; cloudscraper handles anti-bot checks.
        scraper = _create_scraper()
        response = scraper.get(
            url, params=params, headers=HEADERS, timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.HTTPError as exc:
        raise RuntimeError(
            f"IDX API returned HTTP {exc.response.status_code} for {url}"
        ) from exc
    except ValueError:
        # Fallback to plain requests when response body is not valid JSON from scraper.
        response = requests.get(
            url, params=params, headers=HEADERS, timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.RequestException as exc:
        raise RuntimeError(f"Request to IDX API failed: {exc}") from exc


def _download_file(url: str) -> bytes:
    try:
        scraper = _create_scraper()
        response = scraper.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.content
    except requests.RequestException as exc:
        raise RuntimeError(f"Failed to download IDX attachment: {exc}") from exc


def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(BytesIO(content))
    text_parts = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            text_parts.append(page_text)
    return "\n".join(text_parts).strip()


def _extract_xlsx_text(content: bytes) -> str:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        rows.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            clean_cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if clean_cells:
                rows.append(" | ".join(clean_cells))
    wb.close()
    return "\n".join(rows).strip()


def _extract_xls_text(content: bytes) -> str:
    wb = xlrd.open_workbook(file_contents=content)
    rows = []
    for sheet in wb.sheets():
        rows.append(f"[Sheet: {sheet.name}]")
        for i in range(sheet.nrows):
            values = sheet.row_values(i)
            clean_cells = [str(cell).strip() for cell in values if str(cell).strip()]
            if clean_cells:
                rows.append(" | ".join(clean_cells))
    return "\n".join(rows).strip()


def _extract_attachment_text(file_name: str, content: bytes) -> str:
    lower_name = file_name.lower()
    if lower_name.endswith(".pdf"):
        return _extract_pdf_text(content)
    if lower_name.endswith(".xlsx"):
        return _extract_xlsx_text(content)
    if lower_name.endswith(".xls"):
        return _extract_xls_text(content)
    return ""


def _focus_financial_text(raw_text: str) -> str:
    lines = [line.strip() for line in (raw_text or "").splitlines() if line.strip()]
    if not lines:
        return ""

    selected_indexes = set()
    for i, line in enumerate(lines):
        lower_line = line.lower()

        # Financial metrics: harus ada keyword DAN angka di baris yang sama
        has_metric_keyword = any(keyword in lower_line for keyword in FINANCIAL_TEXT_KEYWORDS)
        has_numeric_value = bool(NUMERIC_ROW_RE.search(line))
        if has_metric_keyword and has_numeric_value:
            selected_indexes.add(i)
            for offset in (-1, 1):
                j = i + offset
                if 0 <= j < len(lines):
                    selected_indexes.add(j)

        # Shareholder: cukup ada keyword saja, ambil konteks +-8 baris
        # karena nama pemegang saham dan angka sering di baris yang berbeda
        has_shareholder_keyword = any(keyword in lower_line for keyword in SHAREHOLDER_TEXT_KEYWORDS)
        if has_shareholder_keyword:
            for offset in range(-8, 9):
                j = i + offset
                if 0 <= j < len(lines):
                    selected_indexes.add(j)

    if not selected_indexes:
        return "\n".join(lines[:140])

    focused_lines = [lines[i] for i in sorted(selected_indexes)]
    return "\n".join(focused_lines[:300])

def _collect_report_text(raw_data: dict) -> tuple[str, list[dict]]:
    attachments = raw_data.get("Attachments") or []
    eligible = []

    for item in attachments:
        file_name = str(item.get("File_Name") or "")
        ext = str(item.get("File_Type") or "").lower()
        if not ext and "." in file_name:
            ext = "." + file_name.lower().split(".")[-1]
        if ext in SUPPORTED_EXTENSIONS:
            eligible.append(item)

    def score_attachment(item: dict) -> int:
        file_name = str(item.get("File_Name") or "").lower()
        score = 0
        for kw in PRIORITY_KEYWORDS:
            if kw in file_name:
                score += 3
        for kw in LOW_PRIORITY_KEYWORDS:
            if kw in file_name:
                score -= 2
        return score

    eligible.sort(key=score_attachment, reverse=True)

    parsed_docs = []
    text_chunks = []
    total_chars = 0

    for item in eligible[:MAX_ATTACHMENTS_TO_PARSE]:
        file_name = str(item.get("File_Name") or "unknown")
        file_path = str(item.get("File_Path") or "")
        file_url = file_path if file_path.startswith("http") else f"{BASE_URL}{file_path}"

        try:
            content = _download_file(file_url)
            extracted = _extract_attachment_text(file_name, content)
            extracted = _focus_financial_text(extracted)
        except Exception:
            extracted = ""

        extracted = extracted.strip()
        extracted = extracted[:MAX_CHARS_PER_FILE]

        parsed_docs.append(
            {
                "file_name": file_name,
                "file_type": item.get("File_Type"),
                "file_url": file_url,
                "extracted_chars": len(extracted),
            }
        )

        if extracted:
            remaining = MAX_TOTAL_CHARS - total_chars
            if remaining <= 0:
                break
            excerpt = extracted[:remaining]
            total_chars += len(excerpt)
            text_chunks.append(f"### Dokumen: {file_name}\n{excerpt}")

    return "\n\n".join(text_chunks).strip(), parsed_docs

def _normalized_lookup(raw_data: dict) -> dict:
    lookup = {}
    for key, value in (raw_data or {}).items():
        normalized = "".join(ch for ch in str(key).lower() if ch.isalnum())
        if normalized and normalized not in lookup:
            lookup[normalized] = value
    return lookup


def _pick_field(raw_data: dict, *aliases: str):
    for alias in aliases:
        if alias in raw_data and raw_data.get(alias) not in (None, ""):
            return raw_data.get(alias)

    normalized_lookup = _normalized_lookup(raw_data)
    for alias in aliases:
        normalized = "".join(ch for ch in str(alias).lower() if ch.isalnum())
        value = normalized_lookup.get(normalized)
        if value not in (None, ""):
            return value

    return None


def _fetch_report_results(symbol: str, year: int, periode: str, report_type: str = "rdf") -> list[dict]:
    url = "https://www.idx.co.id/primary/ListedCompany/GetFinancialReport"
    params = {
        "ReportType": report_type,
        "KodeEmiten": symbol.upper(),
        "Year": str(year),
        "SortColumn": "KodeEmiten",
        "SortOrder": "asc",
        "EmitenType": "s",
        "Periode": periode,
        "indexfrom": 1,
        "pagesize": 12,
    }

    response_data = _get(url, params)
    return response_data.get("Results") or []


def get_financial_report(symbol: str, year: int, quarter: str | None = None) -> dict:
    """
    Get financial report from IDX API
    
    Args:
        symbol: Stock code (e.g., 'BBRI', 'ASII')
        year: Year of report (e.g., 2024)
        quarter: Quarter (Q1, Q2, Q3, Q4), optional for yearly mode
    
    Returns:
        dict: Financial report data
    """
    requested_quarter = (quarter or "").strip().upper()
    periode = QUARTER_MAP.get(requested_quarter) if requested_quarter else "audit"

    try:
        # Primary mode from IDX search contract: reportType=rdf and periode by request.
        results = _fetch_report_results(symbol, year, periode, report_type="rdf")
        if results:
            return results[0]

        # Fallback for potential period casing differences.
        fallback_results = _fetch_report_results(symbol, year, periode.upper(), report_type="rdf")
        if fallback_results:
            return fallback_results[0]

        # Compatibility fallback.
        legacy_results = _fetch_report_results(symbol, year, periode, report_type="PDF")
        if legacy_results:
            return legacy_results[0]

        raise RuntimeError(f"No financial data found for {symbol} - {year} {periode}")
            
    except Exception as exc:
        raise RuntimeError(f"Failed to fetch financial report: {exc}") from exc

def parse_financial_data(raw_data: dict) -> dict:
    """Parse raw IDX API response into standardized format"""
    return {
        "kode_emiten": _pick_field(raw_data, "KodeEmiten", "Code"),
        "nama_emiten": _pick_field(raw_data, "NamaEmiten", "Name"),
        "periode_laporan": _pick_field(raw_data, "PeriodeLaporan", "Report_Period", "Period"),
        "tanggal_laporan": _pick_field(raw_data, "TanggalLaporan", "Report_Date", "File_Modified", "Date"),
        "sector": _pick_field(raw_data, "Sector", "Sektor"),
        "sub_sector": _pick_field(raw_data, "SubSector", "Sub_Sector", "SubSektor"),
        "revenue": _pick_field(raw_data, "Revenue", "TotalRevenue", "Sales"),
        "cost_of_goods_sold": _pick_field(raw_data, "CostOfGoodsSold", "COGS"),
        "gross_profit": _pick_field(raw_data, "GrossProfit"),
        "operating_expense": _pick_field(raw_data, "OperatingExpense", "OperatingExpenses"),
        "operating_profit": _pick_field(raw_data, "OperatingProfit", "OperatingIncome"),
        "net_profit": _pick_field(raw_data, "NetProfit", "NetIncome", "ProfitForTheYear", "ProfitLoss"),
        "total_assets": _pick_field(raw_data, "TotalAssets", "TotalAsset", "Assets"),
        "total_liabilities": _pick_field(raw_data, "TotalLiabilities", "Liabilities", "TotalLiability"),
        "total_equity": _pick_field(raw_data, "TotalEquity", "Equity"),
        "eps": _pick_field(raw_data, "EPS", "EarningPerShare"),
        "book_value_per_share": _pick_field(raw_data, "BookValuePerShare", "BVPS"),
        "roe": _pick_field(raw_data, "ROE", "ReturnOnEquity"),
        "roa": _pick_field(raw_data, "ROA", "ReturnOnAssets"),
        "npm": _pick_field(raw_data, "NPM", "NetMargin", "NetProfitMargin"),
        "der": _pick_field(raw_data, "DER", "DebtToEquity"),
        "per": _pick_field(raw_data, "PER", "PriceEarningsRatio"),
        "pbr": _pick_field(raw_data, "PBR", "PriceToBookRatio"),
        "current_ratio": _pick_field(raw_data, "CurrentRatio"),
    }

def scrape_fundamental(symbol: str, year: int, quarter: str | None = None) -> dict:
    """Main function to scrape fundamental data from IDX"""
    raw_data = get_financial_report(symbol, year, quarter)

    parsed_data = parse_financial_data(raw_data)
    report_text, parsed_documents = _collect_report_text(raw_data)
    request_period = (quarter or "").strip().upper() or "AUDIT"

    return {
        "symbol": symbol.upper(),
        "year": year,
        "quarter": request_period,
        "data": parsed_data,
        "report_text": report_text,
        "report_documents": parsed_documents,
        "raw_response": raw_data,
    }

def find_shareholders(data: str) -> list[dict]:
    """
    Extract all shareholders from report text.
    Returns a list of dicts with keys: name, shares, ownership.
    Returns empty list if none found.
    """
    import re

    if not data:
        return []

    print("data length:", len(data))
    text = data

    # Cari semua angka besar (format ribuan dengan titik, misal 1.234.567.890)
    matches = list(re.finditer(r"\d{1,3}(?:\.\d{3}){2,}", text))

    if not matches:
        return []

    print(f"Found {len(matches)} numeric candidates in text.")
    candidates = []
    seen_shares = set()

    for m in matches:
        raw_number = m.group(0)
        shares = int(raw_number.replace(".", ""))

        # Skip angka terlalu kecil (bukan jumlah saham signifikan)
        if shares < 1_000_000:
            continue

        # Deduplicate angka yang sama persis
        if shares in seen_shares:
            continue
        seen_shares.add(shares)

        start = max(0, m.start() - 150)
        end = min(len(text), m.end() + 150)
        context = text[start:end]

        # Cari persentase kepemilikan di sekitar angka
        percent_match = re.search(r"\b\d{1,3}[,\.]\d{2}\b", context)
        ownership = None
        if percent_match:
            try:
                ownership = float(percent_match.group(0).replace(",", "."))
                # Validasi range persentase wajar
                if not (0 < ownership <= 100):
                    ownership = None
            except ValueError:
                ownership = None

        # Ambil nama dari baris sebelum angka
        before_number = context.split(raw_number)[0]
        lines = [l.strip() for l in before_number.strip().split("\n") if l.strip()]
        name = lines[-1] if lines else ""
        name = re.sub(r"\d[\d\.,]*", "", name).strip()
        name = re.sub(r"\s{2,}", " ", name).strip(" |:-")

        if not name:
            continue

        #  FILTER WAJIB: harus ada konteks shareholder
        if not any(k in context.lower() for k in [
            "pemegang saham",
            "shareholder",
            "komposisi pemegang saham"
        ]):
            continue

        # FILTER NOISE (hindari modal/penerbitan)
        if any(bad in context.lower() for bad in [
            "modal",
            "penerbitan",
            "issued",
            "capital",
            "treasury"
        ]):
            continue

        candidates.append({
            "name": name,
            "shares": shares,
            "ownership": ownership,
        })

    # Urutkan dari terbesar ke terkecil
    candidates.sort(key=lambda x: x["shares"], reverse=True)

    return candidates


# Alias untuk backward compatibility
def find_largest_shareholder(data: str):
    shareholders = find_shareholders(data)
    if not shareholders:
        print("No shareholders found in text.")
        return None 
    return shareholders[0]