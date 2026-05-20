from __future__ import annotations

from io import BytesIO

import cloudscraper
import requests
from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd

from config.settings import REQUEST_TIMEOUT

BASE_URL = "https://www.idx.co.id"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": "https://www.idx.co.id/",
}

_IDX_SCRAPER = None


def _create_scraper():
    return cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )


def _get_idx_scraper(reset: bool = False):
    global _IDX_SCRAPER
    if _IDX_SCRAPER is None or reset:
        _IDX_SCRAPER = _create_scraper()
    return _IDX_SCRAPER


def _get_json_cloudscraper(url: str, params: dict | None = None, headers: dict | None = None) -> dict:
    scraper = _get_idx_scraper()
    try:
        response = scraper.get(url, params=params or {}, headers=headers or HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.json()
    except Exception:
        scraper = _get_idx_scraper(reset=True)
        response = scraper.get(url, params=params or {}, headers=headers or HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.json()


def _get(url: str, params: dict, headers: dict | None = None) -> dict:
    try:
        scraper = _create_scraper()
        response = scraper.get(url, params=params, headers=headers or HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.json()
    except requests.HTTPError as exc:
        raise RuntimeError(f"IDX API returned HTTP {exc.response.status_code} for {url}") from exc
    except ValueError:
        response = requests.get(url, params=params, headers=headers or HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as exc:
        raise RuntimeError(f"Request to IDX API failed: {exc}") from exc


def _download_file(url: str) -> bytes:
    try:
        scraper = _create_scraper()
        response = scraper.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.content
    except requests.RequestException as exc:
        raise RuntimeError(f"Failed to download IDX attachment: {exc}") from exc


def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(BytesIO(content))
    text_parts = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            text_parts.append(page_text)
    return "\n".join(text_parts).strip()


def _extract_xlsx_text(content: bytes) -> str:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        rows.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            clean_cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if clean_cells:
                rows.append(" | ".join(clean_cells))
    wb.close()
    return "\n".join(rows).strip()


def _extract_xls_text(content: bytes) -> str:
    wb = xlrd.open_workbook(file_contents=content)
    rows = []
    for sheet in wb.sheets():
        rows.append(f"[Sheet: {sheet.name}]")
        for i in range(sheet.nrows):
            values = sheet.row_values(i)
            clean_cells = [str(cell).strip() for cell in values if str(cell).strip()]
            if clean_cells:
                rows.append(" | ".join(clean_cells))
    return "\n".join(rows).strip()


def _extract_attachment_text(file_name: str, content: bytes) -> str:
    lower_name = file_name.lower()
    if lower_name.endswith(".pdf"):
        return _extract_pdf_text(content)
    if lower_name.endswith(".xlsx"):
        return _extract_xlsx_text(content)
    if lower_name.endswith(".xls"):
        return _extract_xls_text(content)
    return ""
