from io import BytesIO

from openpyxl import Workbook

from app import create_app
from app.scrapers import financial_statement as financial_statement_module


def _build_financial_statement_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    ws.append(["Metric", 2025, 2024])
    ws.append(["Revenue", 1000, 900])
    ws.append(["COGS", 400, 350])
    ws.append(["Gross Profit", 600, 550])
    ws.append(["Operating Expenses", 120, 110])
    ws.append(["Selling Expenses", 40, 35])
    ws.append(["General Admin Expenses", 30, 28])
    ws.append(["R&D Expenses", 10, 9])
    ws.append(["Depreciation & Amortization", 15, 14])
    ws.append(["EBIT", 480, 440])
    ws.append(["EBITDA", 495, 454])
    ws.append(["Operating Income", 470, 430])
    ws.append(["Interest Expense", 20, 18])
    ws.append(["Interest Income", 5, 4])
    ws.append(["Other Non Operating Income", 7, 6])
    ws.append(["Pretax Income", 462, 422])
    ws.append(["Income Tax Expense", 92, 84])
    ws.append(["Effective Tax Rate", 19.91, 19.91])
    ws.append(["Net Income", 370, 338])
    ws.append(["Net Income Attributable", 360, 330])
    ws.append(["Minority Interest", 10, 8])
    ws.append(["EPS", 12.5, 11.2])
    ws.append(["EPS Diluted", 12.2, 11.0])
    ws.append(["Weighted Avg Shares", 300, 302])

    bs = wb.create_sheet("Balance Sheet")
    bs.append(["Metric", 2025, 2024])
    bs.append(["Cash and Cash Equivalents", 200, 180])
    bs.append(["Accounts Receivable", 120, 115])
    bs.append(["Inventory", 80, 70])
    bs.append(["Total Current Assets", 500, 460])
    bs.append(["Property Plant Equipment", 700, 680])
    bs.append(["Intangible Assets", 50, 45])
    bs.append(["Total Non Current Assets", 900, 870])
    bs.append(["Total Assets", 1400, 1330])
    bs.append(["Short Term Debt", 90, 88])
    bs.append(["Accounts Payable", 110, 102])
    bs.append(["Total Current Liabilities", 300, 280])
    bs.append(["Long Term Debt", 250, 245])
    bs.append(["Total Non Current Liabilities", 300, 295])
    bs.append(["Total Liabilities", 600, 575])
    bs.append(["Common Stock", 100, 100])
    bs.append(["Additional Paid In Capital", 50, 50])
    bs.append(["Retained Earnings", 620, 585])
    bs.append(["Total Equity", 800, 755])

    cf = wb.create_sheet("Cash Flow Statement")
    cf.append(["Metric", 2025, 2024])
    cf.append(["Net Income", 500, 450])
    cf.append(["Depreciation & Amortization", 60, 55])
    cf.append(["Stock Based Compensation", 12, 10])
    cf.append(["Change in Working Capital", -5, -4])
    cf.append(["Change in Receivables", -10, -9])
    cf.append(["Change in Inventory", -15, -12])
    cf.append(["Change in Payables", 20, 17])
    cf.append(["Other Operating Activities", 8, 7])
    cf.append(["Net Cash From Operations", 570, 514])
    cf.append(["Capital Expenditures", -80, -70])
    cf.append(["Acquisitions", -20, 0])
    cf.append(["Purchase Of Investments", -40, -30])
    cf.append(["Sale Of Investments", 10, 8])
    cf.append(["Other Investing Activities", 5, 4])
    cf.append(["Net Cash From Investing", -125, -88])
    cf.append(["Debt Issuance", 100, 90])
    cf.append(["Debt Repayment", -60, -55])
    cf.append(["Common Stock Issuance", 15, 12])
    cf.append(["Common Stock Repurchase", -5, -4])
    cf.append(["Dividends Paid", -30, -25])
    cf.append(["Other Financing Activities", 2, 1])
    cf.append(["Net Cash From Financing", 22, 19])
    cf.append(["Net Change In Cash", 467, 445])
    cf.append(["Cash Beginning Period", 100, 80])
    cf.append(["Cash End Period", 567, 525])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_financial_statement_route_requires_symbol_and_year():
    app = create_app()
    client = app.test_client()

    response = client.get("/api/financial-statement")

    assert response.status_code == 400
    body = response.get_json()
    assert body["status"] == "error"
    assert any("symbol" in error for error in body["errors"])
    assert any("year" in error for error in body["errors"])


def test_financial_statement_route_returns_payload(monkeypatch):
    def fake_fetch_and_build_financial_statement(symbol, year):
        return {
            "status": "ok",
            "symbol": symbol,
            "year": year,
            "income_statement": {
                "count": 1,
                "items": [
                    {
                        "period": "AUDIT",
                        "fiscalYear": year,
                        "revenue": 1000,
                    }
                ],
            },
            "balance_sheet": {
                "count": 1,
                "items": [
                    {
                        "period": "AUDIT",
                        "fiscalYear": year,
                        "totalAssets": 1400,
                    }
                ],
            },
            "cash_flow_statement": {
                "count": 1,
                "items": [
                    {
                        "period": "AUDIT",
                        "fiscalYear": year,
                        "netCashFromOperations": 570,
                    }
                ],
            },
        }

    monkeypatch.setattr("app.routes.fetch_and_build_financial_statement", fake_fetch_and_build_financial_statement)

    app = create_app()
    client = app.test_client()

    response = client.get("/api/financial-statement?symbol=bbri&year=2025")

    assert response.status_code == 200
    body = response.get_json()
    assert body["symbol"] == "BBRI"
    assert body["income_statement"]["count"] == 1
    assert body["income_statement"]["items"][0]["revenue"] == 1000
    assert body["balance_sheet"]["count"] == 1
    assert body["balance_sheet"]["items"][0]["totalAssets"] == 1400
    assert body["cash_flow_statement"]["count"] == 1
    assert body["cash_flow_statement"]["items"][0]["netCashFromOperations"] == 570


def test_scrape_financial_statement_extracts_values(monkeypatch):
    workbook_bytes = _build_financial_statement_workbook()

    monkeypatch.setattr(financial_statement_module, "fetch_financial_report_results", lambda symbol, year: [
        {
            "Report_Period": "Audit",
            "Report_Year": str(year),
            "Attachments": [
                {
                    "File_Name": "FinancialStatement-2025-Tahunan-BBRI.xlsx",
                    "File_Path": "/fake/report.xlsx",
                    "File_Type": ".xlsx",
                }
            ],
        }
    ])
    monkeypatch.setattr(financial_statement_module, "_download_file", lambda url: workbook_bytes)

    payload = financial_statement_module.scrape_financial_statement("bbri", 2025)

    assert payload["income_statement"]["count"] == 1
    statement = payload["income_statement"]["items"][0]
    assert statement["revenue"] == 1000
    assert statement["grossProfit"] == 600
    assert statement["netIncome"] == 370
    assert statement["eps"] == 12.5
    assert statement["period"] == "AUDIT"
    assert payload["balance_sheet"]["count"] == 1
    balance_sheet = payload["balance_sheet"]["items"][0]
    assert balance_sheet["totalAssets"] == 1400
    assert balance_sheet["totalLiabilities"] == 600
    assert balance_sheet["totalEquity"] == 800
    assert payload["cash_flow_statement"]["count"] == 1
    cash_flow_statement = payload["cash_flow_statement"]["items"][0]
    assert cash_flow_statement["netIncomeStart"] == 500
    assert cash_flow_statement["depreciationAmort"] == 60
    assert cash_flow_statement["netCashFromOperations"] == 570
    assert cash_flow_statement["freeCashFlow"] == 490


def test_scrape_financial_statement_splits_period_from_file_name(monkeypatch):
    workbook_bytes = _build_financial_statement_workbook()

    monkeypatch.setattr(financial_statement_module, "fetch_financial_report_results", lambda symbol, year: [
        {
            "Report_Period": "Audit",
            "Report_Year": str(year),
            "Attachments": [
                {"File_Name": "FinancialStatement-2025-Tahunan-BBCA.xlsx", "File_Path": "/fake/audit.xlsx", "File_Type": ".xlsx"},
                {"File_Name": "FinancialStatement-2025-I-BBCA.xlsx", "File_Path": "/fake/q1.xlsx", "File_Type": ".xlsx"},
                {"File_Name": "FinancialStatement-2025-II-BBCA.xlsx", "File_Path": "/fake/q2.xlsx", "File_Type": ".xlsx"},
                {"File_Name": "FinancialStatement-2025-III-BBCA.xlsx", "File_Path": "/fake/q3.xlsx", "File_Type": ".xlsx"},
            ],
        }
    ])
    monkeypatch.setattr(financial_statement_module, "_download_file", lambda url: workbook_bytes)

    payload = financial_statement_module.scrape_financial_statement("bbca", 2025)

    income_periods = [item.get("period") for item in payload["income_statement"]["items"]]
    balance_periods = [item.get("period") for item in payload["balance_sheet"]["items"]]
    assert payload["income_statement"]["count"] == 4
    assert payload["balance_sheet"]["count"] == 4
    assert "AUDIT" in income_periods
    assert "Q1" in income_periods
    assert "Q2" in income_periods
    assert "Q3" in income_periods
    assert "AUDIT" in balance_periods
    assert "Q1" in balance_periods
    assert "Q2" in balance_periods
    assert "Q3" in balance_periods
